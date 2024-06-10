import openpyxl


def xpaths_input(ruta_archivo):
    workbook = openpyxl.load_workbook(ruta_archivo)
    sheet = workbook.active
    xpaths = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        xpath = row[0].value
        xpaths.append(xpath)
    return xpaths
