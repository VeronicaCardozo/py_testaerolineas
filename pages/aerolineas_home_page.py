from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
import allure
import time
import os
import openpyxl
import csv

import openpyxl
from datetime import datetime, timedelta


class AerolineasHomePage:

    # Locators
    btn_locator_aceptar_cookies = (
        By.XPATH, "(//button[@type='cookies'])")
    btn_locator_lenguaje_menu = (
        By.XPATH, "//button[@class='styled__LanguageMenu-sc-22nvvu-8 fEfYIU']")
    lenguaje_elegido = (
        By.XPATH, "//div[@id='language-panel']//a[normalize-space()='United States (English)']")
    box_origen = (By.XPATH, "//input[@placeholder='Origen']")
    box_destino = (By.XPATH, "//input[@placeholder='Destino']")
    box_date_from = (By.XPATH, "//input[@id='from-date']")
    box_date_return = (By.XPATH, "//input[@id='to-date']")
    number_pasajeros = (By.XPATH, "//button[@id = 'cabin-passengers']")
    sumar_pasajeros = (
        By.CSS_SELECTOR, "button.styled__IconContainer-sc-1sy3ra0-1.eoncfD.add-adt")
    btn_click_vuelo = (By. XPATH, "//button[@id= 'search-flights']")

    # Localizadores para los elementos
    fecha_ida = (
        By.XPATH, "//div[contains(@class, 'styled__EnabledDateOffer-ty299w-8')]//div[contains(@class, 'fdc-button-day') and normalize-space()='1']")
    fecha_vuelta = (
        By.XPATH, "//div[contains(@class, 'styled__EnabledDateOffer-ty299w-8')]//div[contains(@class, 'fdc-button-day') and normalize-space()='7']")
    label_locator_monto = (
        By.XPATH, "//div[@class = 'styled__DateOfferItem-ty299w-6 styled__EnabledDateOffer-ty299w-8 guJlAe fdc-available-day']//div[@class = 'styled__Price-ty299w-0 cbwzbP fdc-button-price'][normalize-space()]")
    btn_ver_vuelo = (By.XPATH, "//button[contains(text(),'Ver vuelos')]")
    card_destino_nacional_0 = (By.XPATH, "(//a[@data-posicion='0'])")
    card_destino_nacional_1 = (By.XPATH, "//a[@data-posicion='1']")
    card_destino_nacional_2 = (By.XPATH, "//a[@data-posicion='2']")
    card_destino_nacional_3 = (By.XPATH, "//a[@data-posicion='3']")
    btn_internacionales_card = (
        By.XPATH, "(//button[normalize-space()='Destinos Internacionales'])")
    btn_regionales_card = (
        By.XPATH, "(//button[normalize-space()='Destinos Regionales'])")
    button_whatsapp_locator = (By.ID, "whatsapp_flotante")
    button_chatbot_locator = (By.XPATH, "//body/div[@id='wcx-chat']/button[1]")
    box_chatbot_locator = (
        By.XPATH, "//div[@class='chat-container']//span[@class='input-group-addon']//input[@class='form-control ng-pristine ng-invalid ng-invalid-required ng-touched div-name-with-buttons']")
    iframe_chat_onlyne_locator = (
        By.XPATH, "//body/div[@id='wcx-chat']/div[1]/div[1]/iframe[1]")
    chat_olyne_name_input_locator = (
        By.XPATH, "//div[@id='f1e92af809cc4af596b8e15009725100']//input[@name='name']")

    def __init__(self, driver):
        self.driver = driver
    # metodos
    # 1) Aceptar cookies y verificacion de links

    @allure.step("Verificamos el boton aceptar cookies y la cantidad de links en el home")
    def click_aceptar_cookies(self):
        btn = WebDriverWait(self.driver, 3).until(
            EC.visibility_of_element_located(
                self.btn_locator_aceptar_cookies)
        )
        print(" Boton aceptar cookies visible")
        assert btn.is_displayed(), "El boton de cookies no esta visible"
        btn.click()
        print(" Boton aceptar cookies visible y se hizo click en el ")

    @allure.step("Verificar cuantos Links hay en el home ")
    def cantidad_links_home(self):
        links = self.driver.find_elements(By.TAG_NAME, "a")
        print("El numero de Links que hay en el home de la pagina Aerolineas Argentina es:", len(links))
        for num in links:
            print(num.text)
        assert len(links) > 0, "No hay links en el home"

    @allure.step("Verificar accesibilidad de links importantes")
    def verificar_links(self):

        links = [
            ("Vuelos", "//a[contains(text(),'VUELOS')]"),
            ("Check-in", "//a[contains(text(),'CHECK IN')]"),
            ("Estado de vuelo", "//a[contains(text(),'ESTADO DE VUELO')]"),
            ("Mi reserva", "//a[contains(text(),'MI RESERVA')]")
        ]
        for name, xpath in links:
            allure.dynamic.description(f"Verificando link: {name}")
            assert self.verify_element(By.XPATH, xpath), f"El enlace '{
                name}' no fue encontrado."

    @allure.step("Verificar links")
    def verify_element(self, by, locator):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((by, locator))
            )
            return True
        except:
            return False

    @allure.step("Verificar boton menu idiomas e idioma seleccionado ")
    def bnt_lenguaje_menu(self):
        btn_lenguaje = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.btn_locator_lenguaje_menu)
        )
        if btn_lenguaje.is_displayed():
            btn_lenguaje.click()
            time.sleep(3)
            print("El menu lenguaje es visible")
            assert btn_lenguaje.is_displayed(), "El boton no esta verificado"
            lenguage_option = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.lenguaje_elegido)
            )
            lenguage_option.click()
            time.sleep(5)
            print("Idioma seleccionado")
        else:
            print("El botón de idioma no está visible")
        # datos excel

    @allure.step("Validar los datos ingresados desde un archivo plano: en este caso Excel")
    def datos_excel(self):
        archivo = openpyxl.load_workbook(
            'C:\\py_automation\\Data\\datos_qa.xlsx')
        sheet = archivo['Hoja1']
        for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            valores_fila = [celda.value for celda in fila]
            origen, destino, ida, regreso, url = valores_fila
            ida_str = ida.strftime("%d/%m/%Y")
            regreso_str = regreso.strftime("%d/%m/%Y")
            self.buscar_vuelo(origen, destino, ida_str, regreso_str)

    @allure.step("Validar los datos de la reserva")
    def buscar_vuelo(self, origen, destino, ida_str, regreso_str):
        box_origen = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_origen))
        box_origen.clear()
        box_origen.send_keys(origen)
        time.sleep(2)
        box_origen.send_keys(Keys.ARROW_DOWN, Keys.ENTER, Keys.TAB)
        print("Caja de origen es visible y los datos son ingresados")
        box_destino = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_destino))
        box_destino.clear()
        box_destino.send_keys(destino)
        time.sleep(2)
        box_destino.send_keys(Keys.ARROW_DOWN, Keys.ENTER, Keys.TAB)
        print("Caja de destino es visible y los datos son ingresados")
        box_date_from = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_date_from))
        box_date_from.clear()
        box_date_from.send_keys(ida_str, Keys.TAB)
        print("Caja de date_from es visible y los datos son ingresados")
        box_date_return = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_date_return))
        box_date_return.clear()
        box_date_return.send_keys(regreso_str, Keys.TAB)
        print("Caja de date_return es visible y los datos son ingresados")

    @allure.step("Validar cantidad de pasajeros=2")
    def cambiar_cantidad_pasajeros(self):
        try:
            pasajeros = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.number_pasajeros))
            pasajeros.click()
            assert pasajeros.is_displayed(), "El boton no esta verificado"
            print("Boton pasajeros visible y hacemos click")
            btn_mas = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.sumar_pasajeros))
            btn_mas.click()
            time.sleep(2)
            cantidad_pasajeros = pasajeros.get_attribute("value")
            if cantidad_pasajeros != "2":
                print("La cantidad de pasajeros se cambió")
            return True
        except TimeoutException as e:
            print(f"Error: {e}")

    @allure.step("Validar la búsqueda ingresando al botón búsqueda de vuelo de un vuelo-ida-regreso")
    def comprar_vuelo(self):
        try:
            btn_comprar = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.btn_click_vuelo))
            if btn_comprar.is_enabled():
                btn_comprar.click()
                time.sleep(6)

                print("El botón de comprar de vuelo es visible y hacemos click en él")
            else:
                print("El botón de compra de vuelo no es visible.")
        except TimeoutException:
            print("No se encontró el botón de comprar de vuelo ")

    @allure.step("Validar los elementos de vuelo y precios en la grilla de vuelos")
    def precios_vuelos(self):
        try:
            vuelos_ida = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_ida))
            time.sleep(3)
            vuelos_vuelta = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_vuelta))
            time.sleep(3)

            print("Vuelos de ida:")
            for vuelo in vuelos_ida:
                dia_vuelo_ida = vuelo.text
                precio_vuelo_ida = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_ida > 0:
                    print("Día:", dia_vuelo_ida, "Precio:", precio_vuelo_ida)

            print("Vuelos de regreso:")
            for vuelo in vuelos_vuelta:
                dia_vuelo_regreso = vuelo.text
                precio_vuelo_regreso = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_regreso > 0:
                    print("Día:", dia_vuelo_regreso,
                          "Precio:", precio_vuelo_regreso)
        except TimeoutException:
            print("Error: No se encontraron vuelos en la fecha seleccionada")

    @allure.step("Validar precios")
    def obtener_precio_vuelo(self, vuelo_elemento):
        try:
            precio_elemento = vuelo_elemento.find_element(
                *self.label_locator_monto)
            precio_text = precio_elemento.text.replace(
                '$', '').replace(',', '').strip()
            if precio_text and precio_text.replace('.', '', 1).isdigit():
                return float(precio_text)
            return 0.0
        except Exception as e:
            print(f"Error al obtener el precio del vuelo: {e}")
            return 0.0

    @allure.step("Validar la búsqueda ingresando al botón búsqueda de vuelo de un vuelo-ida-regreso")
    def comprar_vuelo(self):
        try:
            btn_comprar = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.btn_click_vuelo))
            if btn_comprar.is_enabled():
                btn_comprar.click()
                time.sleep(6)

                print("El botón de comprar de vuelo es visible y hacemos click en él")
            else:
                print("El botón de compra de vuelo no es visible.")
        except TimeoutException:
            print("No se encontró el botón de comprar de vuelo ")

    @allure.step("Validar los elementos de vuelo y precios en la grilla de vuelos")
    def precios_vuelos(self):
        try:
            vuelos_ida = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_ida))
            time.sleep(3)
            vuelos_vuelta = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_vuelta))
            time.sleep(3)

            print("Vuelos de ida:")
            for vuelo in vuelos_ida:
                dia_vuelo_ida = vuelo.text
                precio_vuelo_ida = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_ida > 0:
                    print("Día:", dia_vuelo_ida, "Precio:", precio_vuelo_ida)

            print("Vuelos de regreso:")
            for vuelo in vuelos_vuelta:
                dia_vuelo_regreso = vuelo.text
                precio_vuelo_regreso = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_regreso > 0:
                    print("Día:", dia_vuelo_regreso,
                          "Precio:", precio_vuelo_regreso)
        except TimeoutException:
            print("Error: No se encontraron vuelos en la fecha seleccionada")

    @allure.step("Validar precios")
    def obtener_precio_vuelo(self, vuelo_elemento):
        try:
            precio_elemento = vuelo_elemento.find_element(
                *self.label_locator_monto)
            precio_text = precio_elemento.text.replace(
                '$', '').replace(',', '').strip()
            if precio_text and precio_text.replace('.', '', 1).isdigit():
                return float(precio_text)
            return 0.0
        except Exception as e:
            print(f"Error al obtener el precio del vuelo: {e}")
            return 0.0

    @allure.step("Validar la compra ingresando al botón ver vuelos")
    def ver_vuelos(self):
        try:
            print("Esperando que el botón de ver vuelos sea visible...")
            ver_vuelo = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.btn_ver_vuelo)
            )

            if ver_vuelo.is_displayed():
                print("El botón de ver vuelos es visible.")
            if ver_vuelo.is_enabled():
                ver_vuelo.click()
                time.sleep(6)
                print("Hicimos click en el botón de ver vuelos.")

                time.sleep(3)
                print(
                    "El botón de ver vuelos es visible y hacemos click en el")
            else:
                print("El botón de ver vuelos no es visible.")
        except TimeoutException:
            print("No se encontró el botón de ver vuelos después de esperar 20 segundos.")

            print(
                "No se encontró el botón de ver vuelos ")

    # test +20 dias en hoja2
    @allure.step("Validar los datos ingresados desde un archivo plano sumando 20 dias a hoja 2")
    def datos_excel_sumar_20(self):
        archivo = openpyxl.load_workbook(
            'C:\\py_automation\\py_testaerolineas\\Data\\datos_qa.xlsx')
        sheet = archivo['Hoja2']
        for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            valores_fila = [celda.value for celda in fila]
            origen, destino, ida, regreso, url = valores_fila
            # Sumar 20 días a la fecha de ida y regreso
            ida = self.sumar_20_dias(ida)
            regreso = self.sumar_20_dias(regreso)
            ida_str = ida.strftime("%d/%m/%Y")
            regreso_str = regreso.strftime("%d/%m/%Y")
            self.buscar_vuelo(origen, destino, ida_str, regreso_str)

    def sumar_20_dias(self, fecha_actual):
        fecha_sumada = fecha_actual + timedelta(days=20)
        return fecha_sumada

    # Test card Nacional

    @allure.step("Validar si se encuentran las Cards de Destinos Nacionales")
    def validate_destino_nacional_card(self):
        card_nacional = WebDriverWait(self.driver, 20).until(
            EC.visibility_of_element_located(self.card_destino_nacional_2))
        actions = ActionChains(self.driver)
        actions.move_to_element(card_nacional).perform()
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Screenshot_destino_nacional", attachment_type=allure.attachment_type.PNG)
        try:
            card_element_0 = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.card_destino_nacional_0))
            if card_element_0.is_displayed():
                return "Se muestra las cards Nacionales"
            else:
                return "No se muestra la cards"
        except Exception as e:
            return "Error Card 1"

    @allure.step("Validar si se encuentran las Cards de Destinos Internacionales")
    def validate_destino_internacional_card(self):
        try:
            internacional = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.btn_international_card))
            internacional.click()
            card_internacional = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.card_destino_nacional_2))
            actions = ActionChains(self.driver)
            actions.move_to_element(card_internacional).perform()
            allure.attach(self.driver.get_screenshot_as_png(
            ), name="Screenshot_destino_internacional", attachment_type=allure.attachment_type.PNG)
            card_element_0 = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.card_destino_nacional_0))
            if card_element_0.is_displayed():
                return "Se muestran las cards Internacionales"
            else:
                return "No se muestran las cards"
        except Exception as e:
            allure.attach(self.driver.get_screenshot_as_png(
            ), name="Error_Screenshot", attachment_type=allure.attachment_type.PNG)
            return f"Error Card Internacional: {str(e)}"

    @allure.step("Validar si se encuentran las Cards de Destinos Regionales")
    def validate_destino_regional_card(self):
        try:
            regional = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.btn_regionales_card))
            regional.click()
            card_regional = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.card_destino_nacional_2))
            actions = ActionChains(self.driver)
            actions.move_to_element(card_regional).perform()
            allure.attach(self.driver.get_screenshot_as_png(
            ), name="Screenshot_destino_regional", attachment_type=allure.attachment_type.PNG)
            card_element_0 = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.card_destino_nacional_0))
            if card_element_0.is_displayed():
                return "Se muestran las cards Regionales"
            else:
                return "No se muestran las cards"
        except Exception as e:
            allure.attach(self.driver.get_screenshot_as_png(
            ), name="Error_Screenshot", attachment_type=allure.attachment_type.PNG)
            return f"Error Card Regional: {str(e)}"

    def validate_card_nacional_csv(self, output_file='ofertas_nacionales.csv'):
        xpaths = [
            "(//a[@data-posicion='0'])",
            "(//a[@data-posicion='1'])",
            "(//a[@data-posicion='2'])",
            "(//a[@data-posicion='3'])"
        ]
        data = []
        try:
            for xpath in xpaths:
                card = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )
                desde = card.find_element(
                    By.XPATH, ".//div[@class='styled__SecondaryLabel-sc-17je9ro-5 tfyso thumbnail-secondary-label']").text
                hacia = card.find_element(
                    By.XPATH, ".//label[@class='styled__PrimaryLabel-sc-17je9ro-4 klasRx thumbnail-primary-label']").text
                precio = card.find_element(
                    By.XPATH, ".//label[@class='styled__Fare-sc-17je9ro-7 fYFkWK thumbnail-fare']").text
                clase = card.find_element(
                    By.XPATH, ".//label[@class='styled__Badge-sc-17je9ro-3 nRgyB thumbnail-badge']").text

                data.append([desde, hacia, clase, precio])

        finally:
            with open(output_file, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['desde', 'hacia', 'clase', 'precio'])
                writer.writerows(data)

            print(f'Data has been written to {output_file}')

    def validate_card_internacional_csv(self, output_file='ofertas_internacionales.csv'):
        xpaths = [
            "(//a[@data-posicion='0'])",
            "(//a[@data-posicion='1'])",
            "(//a[@data-posicion='2'])",
            "(//a[@data-posicion='3'])"
        ]
        data = []
        try:
            for xpath in xpaths:
                card = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )
                desde = card.find_element(
                    By.XPATH, ".//div[@class='styled__SecondaryLabel-sc-17je9ro-5 tfyso thumbnail-secondary-label']").text
                hacia = card.find_element(
                    By.XPATH, ".//label[@class='styled__PrimaryLabel-sc-17je9ro-4 klasRx thumbnail-primary-label']").text
                precio = card.find_element(
                    By.XPATH, ".//label[@class='styled__Fare-sc-17je9ro-7 fYFkWK thumbnail-fare']").text
                clase = card.find_element(
                    By.XPATH, ".//label[@class='styled__Badge-sc-17je9ro-3 nRgyB thumbnail-badge']").text

                data.append([desde, hacia, clase, precio])

        finally:
            with open(output_file, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['desde', 'hacia', 'clase', 'precio'])
                writer.writerows(data)

            print(f'Data has been written to {output_file}')

    def validate_card_regional_csv(self, output_file='ofertas_regionales.csv'):
        xpaths = [
            "(//a[@data-posicion='0'])",
            "(//a[@data-posicion='1'])",
            "(//a[@data-posicion='2'])",
            "(//a[@data-posicion='3'])"
        ]
        data = []
        try:
            for xpath in xpaths:
                card = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )
                desde = card.find_element(
                    By.XPATH, ".//div[@class='styled__SecondaryLabel-sc-17je9ro-5 tfyso thumbnail-secondary-label']").text
                hacia = card.find_element(
                    By.XPATH, ".//label[@class='styled__PrimaryLabel-sc-17je9ro-4 klasRx thumbnail-primary-label']").text
                precio = card.find_element(
                    By.XPATH, ".//label[@class='styled__Fare-sc-17je9ro-7 fYFkWK thumbnail-fare']").text
                clase = card.find_element(
                    By.XPATH, ".//label[@class='styled__Badge-sc-17je9ro-3 nRgyB thumbnail-badge']").text

                data.append([desde, hacia, clase, precio])

        finally:
            with open(output_file, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['desde', 'hacia', 'clase', 'precio'])
                writer.writerows(data)

            print(f'Data has been written to {output_file}')
<<<<<<< HEAD

    @allure.step("Verificación whatsapp web")
    def whatsapp_chat(self):
        btn_wts = WebDriverWait(self.driver, 20).until(
            EC.visibility_of_element_located(self.button_whatsapp_locator)
        )
        assert btn_wts.is_displayed(), "El botón de WhatsApp no es visible"
        btn_wts.click()
        print("Se hizo clic en el botón de WhatsApp")

        # Esperar a que aparezca la ventana emergente de WhatsApp
        WebDriverWait(self.driver, 20).until(
            lambda driver: len(driver.window_handles) > 1
        )

        # Cambiar al nuevo handle (ventana emergente)
        new_handle = self.driver.window_handles[-1]
        self.driver.switch_to.window(new_handle)

        # Verificar el enlace de la ventana emergente de WhatsApp
        current_url = self.driver.current_url
        assert "https://api.whatsapp.com/send?phone=541149404798" in current_url, f"La URL actual no es de WhatsApp: {
            current_url}"

        print("Verificación de WhatsApp completada correctamente")

# En tu página de inicio de Aerolíneas, verifica el botón de chat online después de hacer clic en WhatsApp

    @allure.step("Verificación chat-online login")
    def login_chat_online(self):
        # Esperar a que el botón del chatbot sea visible
        btn_chat_online = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.button_chatbot_locator)
        )
        assert btn_chat_online.is_displayed(), "El botón de chat no es visible"
        btn_chat_online.click()
        print("Se hizo click en el botón")

        # Cambiar al iframe del chat en línea
        WebDriverWait(self.driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it(
                self.iframe_chat_onlyne_locator)
        )
        print("Se cambió al iframe del chat en línea")

        # Ingresar un nombre en el campo de entrada name
        input_locator_name = (By.NAME, "name")
        input_element_name = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_name)
        )
        input_element_name.send_keys("Nombre de prueba")

        # Ingresar un correo en el campo de entrada email
        input_locator_email = (By.NAME, "email")
        input_element_email = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_email)
        )
        input_element_email.send_keys("python@automation.com")

        # Ingresar información en el campo de entrada pasajero_frecuente
        input_locator_pasajero_frecuente = (By.NAME, "pasajero_frecuente")
        input_element_pasajero_frecuente = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_pasajero_frecuente)
        )
        input_element_pasajero_frecuente.send_keys("no tengo")

        # Hacer click en el botón "Continuar"
        continuar_button_locator = (
            By.XPATH, "//input[@type='submit' and @value='Continuar']")
        button_element_continuar = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(continuar_button_locator)
        )
        button_element_continuar.click()

        # Verificar que se haya cambiado al nuevo contexto después de hacer click en Continuar
        WebDriverWait(self.driver, 10).until(
            EC.staleness_of(button_element_continuar))
        print("Se cambió de contexto después de hacer click en Continuar")

        # Volver al contexto padre si es necesario
        self.driver.switch_to.default_content()
        print("Verificación de chat online completada correctamente")

    @allure.step("Verificación chat-online boton mis reservas")
    def reservas_chat_online(self):
        self.login_chat_online()
        # Localizar y hacer click en el botón "Ver mi reserva"
        button_reserva_locator = (
            By.XPATH, "//button[contains(text(), 'Ver mi reserva')]")
        button_element_reserva = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(button_reserva_locator)
        )
        assert button_element_reserva.is_displayed(
        ), "El botón 'Ver mi reserva' no es visible"
        button_element_reserva.click()
        print("Se hizo click en el botón 'Ver mi reserva'")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Botón Ver mi reserva clickeado", attachment_type=allure.attachment_type.PNG)

        # Ingresar un número de reserva incorrecto
        input_locator_reserva = (By.NAME, "chat")
        input_element_reserva = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_reserva)
        )
        # Número de reserva incorrecto
        input_element_reserva.send_keys("1234567")
        print("se ingresa numero incorrecto de reserva")

        # Enviar el número de reserva incorrecto
        enviar_button_locator = (
            By.ID, "send")
        button_element_enviar = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(enviar_button_locator)
        )
        button_element_enviar.click()

        print("Se verifico que el chat responde'")
        # Volver al contexto padre si es necesario
        self.driver.switch_to.default_content()

        print("Verificación del boton reservas en chat online completada correctamente")

    @allure.step("Verificación chat-online No valide mis reservas")
    def reservas_chat_online2(self):
        # Esperar a que el botón del chatbot sea visible
        btn_chat_online = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.button_chatbot_locator)
        )
        assert btn_chat_online.is_displayed(), "El botón de chat no es visible"
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Botón chatbot visible", attachment_type=allure.attachment_type.PNG)
        btn_chat_online.click()
        print("Se hizo click en el botón")

        # Cambiar al iframe del chat en línea
        iframe_chat_online = WebDriverWait(self.driver, 10).until(
            EC.frame_to_be_available_and_switch_to_it(
                self.iframe_chat_onlyne_locator)
        )
        print("Se cambió al iframe del chat en línea")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Cambio a iframe del chat", attachment_type=allure.attachment_type.PNG)

        # Ingresar un nombre en el campo de entrada name
        input_locator_name = (By.NAME, "name")
        input_element_name = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_name)
        )
        input_element_name.send_keys("Nombre de prueba")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Nombre ingresado", attachment_type=allure.attachment_type.PNG)

        # Ingresar un correo en el campo de entrada email
        input_locator_email = (By.NAME, "email")
        input_element_email = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_email)
        )
        input_element_email.send_keys("python@automation.com")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Email ingresado", attachment_type=allure.attachment_type.PNG)

        # Ingresar información en el campo de entrada pasajero_frecuente
        input_locator_pasajero_frecuente = (By.NAME, "pasajero_frecuente")
        input_element_pasajero_frecuente = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_pasajero_frecuente)
        )
        input_element_pasajero_frecuente.send_keys("no tengo")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Información ingresada", attachment_type=allure.attachment_type.PNG)

        # Hacer click en el botón "Continuar"
        continuar_button_locator = (
            By.XPATH, "//input[@type='submit' and @value='Continuar']")
        button_element_continuar = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(continuar_button_locator)
        )
        button_element_continuar.click()
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Botón Continuar clickeado", attachment_type=allure.attachment_type.PNG)

        # Verificar que se haya cambiado al nuevo contexto después de hacer click en Continuar
        WebDriverWait(self.driver, 10).until(
            EC.staleness_of(button_element_continuar))
        print("Se cambió de contexto después de hacer click en Continuar")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Cambio de contexto verificado", attachment_type=allure.attachment_type.PNG)

        # Localizar y hacer click en el botón "Ver mi reserva"
        button_reserva_locator = (
            By.XPATH, "//button[contains(text(), 'Ver mi reserva')]")
        button_element_reserva = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(button_reserva_locator)
        )
        assert button_element_reserva.is_displayed(
        ), "El botón 'Ver mi reserva' no es visible"
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Botón Ver mi reserva visible", attachment_type=allure.attachment_type.PNG)
        button_element_reserva.click()
        print("Se hizo click en el botón 'Ver mi reserva'")
        allure.attach(self.driver.get_screenshot_as_png(
        ), name="Botón Ver mi reserva clickeado", attachment_type=allure.attachment_type.PNG)

        # Ingresar un número de reserva incorrecto
        input_locator_reserva = (By.NAME, "chat")
        input_element_reserva = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(input_locator_reserva)
        )
        # Número de reserva incorrecto
        input_element_reserva.send_keys("1234567")
        print("se ingresa numero incorrecto de reserva")

        # Enviar el número de reserva incorrecto
        enviar_button_locator = (
            By.ID, "send")
        button_element_enviar = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(enviar_button_locator)
        )
        button_element_enviar.click()

        print("Se verifico que el chat responde'")
        # Volver al contexto padre si es necesario
        self.driver.switch_to.default_content()
=======
    btn_locator_aceptar_cookies = (By.XPATH, "//button[@id='cookies']")
    btn_locator_vuelos = (By.XPATH, "//a[normalize-space()='VUELOS']")
    btn_locator_check_in = (By.XPATH, "//a[normalize-space()='CHECK IN']")

    @allure.step("Hacemos click en el boton aceptar cookies")
    def clic_aceptar_cookies(self):
        """
        Método para validar el botón aceptar cookies.
        """
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.btn_locator_aceptar_cookies)
        )
        if self.driver.find_element(*self.btn_locator_aceptar_cookies).is_displayed():
            self.driver.find_element(*self.btn_locator_aceptar_cookies).click()

    @allure.step("Verificar el boton vuelos")
    def btn_vuelos(self):
        """
        Método para validar el boton vuelo.
        """
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.btn_locator_vuelos)
        )
        if self.driver.find_element(*self.btn_locator_vuelos).is_displayed():
            self.driver.find_element(*self.btn_locator_vuelos).click()

    @allure.step("Verificar el boton chek in")
    def btn_check_in(self):
        """
        Método para validar el boton check in.
        """
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.btn_locator_check_in)
        )
        if self.driver.find_element(*self.btn_locator_check_in).is_displayed():
            self.driver.find_element(*self.btn_locator_check_in).click()
>>>>>>> 975f161ac7ffea0997d062bdc5ad4b2d2ed253a5
