from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import ElementClickInterceptedException
import allure
import time
import os
import openpyxl
import csv

import openpyxl
from datetime import datetime, timedelta


class home_page_vero:

    # Locators
    btn_locator_aceptar_cookies = (
        By.XPATH, "(//button[@type='cookies'])")
    btn_locator_lenguaje_menu = (
        By.XPATH, "//button[@class='styled__LanguageMenu-sc-22nvvu-8 fEfYIU']")
    lenguaje_elegido = (
        By.XPATH, "//div[@id='language-panel']//a[normalize-space()='United States (English)']")
    box_origen = (By.XPATH, "//input[@placeholder='Origen']")
    box_destino = (By.XPATH, "//input[@placeholder='Destino']")
    box_date_from = (By.XPATH, "//input[@id='from-date']")
    box_date_return = (By.XPATH, "//input[@id='to-date']")
    number_pasajeros = (By.XPATH, "//button[@id = 'cabin-passengers']")
    sumar_pasajeros = (
        By.CSS_SELECTOR, "button.styled__IconContainer-sc-1sy3ra0-1.eoncfD.add-adt")
    btn_click_vuelo = (By. XPATH, "//button[@id= 'search-flights']")

    # Localizadores para los elementos
    fecha_ida = (
        By.XPATH, "//div[contains(@class, 'styled__EnabledDateOffer-ty299w-8')]//div[contains(@class, 'fdc-button-day') and normalize-space()='1']")
    fecha_vuelta = (
        By.XPATH, "//div[contains(@class, 'styled__EnabledDateOffer-ty299w-8')]//div[contains(@class, 'fdc-button-day') and normalize-space()='7']")
    label_locator_monto = (
        By.XPATH, "//div[@class = 'styled__DateOfferItem-ty299w-6 styled__EnabledDateOffer-ty299w-8 guJlAe fdc-available-day']//div[@class = 'styled__Price-ty299w-0 cbwzbP fdc-button-price'][normalize-space()]")
    btn_ver_vuelo = (By.XPATH, "//button[contains(text(),'Ver vuelos')]")

    def __init__(self, driver):
        self.driver = driver
    # metodos
    # 1) Aceptar cookies y verificacion de links

    @allure.step("Verificamos el boton aceptar cookies y la cantidad de links en el home")
    def click_aceptar_cookies(self):
        btn = WebDriverWait(self.driver, 3).until(#esperas explicitas
            EC.visibility_of_element_located(
                self.btn_locator_aceptar_cookies)
        )
        print(" Boton aceptar cookies visible")
        assert btn.is_displayed(), "El boton de cookies no esta visible"
        btn.click()
        print(" Boton aceptar cookies visible y se hizo click en el ")
    #cantidad de links home
    @allure.step("Verificar cuantos Links hay en el home ")
    def cantidad_links_home(self):
        links = self.driver.find_elements(By.TAG_NAME, "a")
        print("El numero de Links que hay en el home de la pagina Aerolineas Argentina es:", len(links))
        for num in links:
            print(num.text)
        assert len(links) > 0, "No hay links en el home"
    #verifica links importantes y me da el resultado en el reporte
    @allure.step("Verificar accesibilidad de links importantes")
    def verificar_links(self):
        links = [
            ("Vuelos", "//a[contains(text(),'VUELOS')]"),
            ("Check-in", "//a[contains(text(),'CHECK IN')]"),
            ("Estado de vuelo", "//a[contains(text(),'ESTADO DE VUELO')]"),
            ("Mi reserva", "//a[contains(text(),'MI RESERVA')]")
        ]
        for name, xpath in links:
            allure.dynamic.description(f"Verificando link: {name}")
            assert self.verify_element(By.XPATH, xpath), f"El enlace '{
                name}' no fue encontrado."
        print("Los links fueron localizados y enviados al reporte de allure")

    @allure.step("Verificar links")
    def verify_element(self, by, locator):
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((by, locator))
            )
            return True
        except:
            return False
    #verifica lenguaje menu y seleciona un idioma(Ingles)
    @allure.step("Verificar boton menu idiomas e idioma seleccionado ")
    def bnt_lenguaje_menu(self):
        btn_lenguaje = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.btn_locator_lenguaje_menu)
        )
        if btn_lenguaje.is_displayed():
            btn_lenguaje.click()
            time.sleep(3)
            print("El menu lenguaje es visible")
            assert btn_lenguaje.is_displayed(), "El boton no esta verificado"
            lenguage_option = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.lenguaje_elegido)
            )
            lenguage_option.click()
            time.sleep(5)
            print("Idioma seleccionado")
        else:
            print("El botón de idioma no está visible")
        
        # datos excel
    @allure.step("Validar los datos ingresados desde un archivo plano: en este caso Excel")
    def datos_excel(self):
        archivo = openpyxl.load_workbook(
            'C:\\py_automation\\Data\\datos_qa.xlsx')
        sheet = archivo['Hoja1']
        for fila in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            valores_fila = [celda.value for celda in fila]
            origen, destino, ida, regreso, url = valores_fila
            ida_str = ida.strftime("%d/%m/%Y")
            regreso_str = regreso.strftime("%d/%m/%Y")
            self.buscar_vuelo(origen, destino, ida_str, regreso_str)
    #introduce datos en las cajas de reservas de vuelos
    @allure.step("Validar los datos de la reserva")
    def buscar_vuelo(self, origen, destino, ida_str, regreso_str):
        box_origen = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_origen))
        box_origen.clear()
        box_origen.send_keys(origen)
        time.sleep(2)
        box_origen.send_keys(Keys.ARROW_DOWN, Keys.ENTER, Keys.TAB)
        print("Caja de origen es visible y los datos son ingresados")
        box_destino = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_destino))
        box_destino.clear()
        box_destino.send_keys(destino)
        time.sleep(2)
        box_destino.send_keys(Keys.ARROW_DOWN, Keys.ENTER, Keys.TAB)
        print("Caja de destino es visible y los datos son ingresados")
        box_date_from = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_date_from))
        box_date_from.clear()
        box_date_from.send_keys(ida_str, Keys.TAB)
        print("Caja de date_from es visible y los datos son ingresados")
        box_date_return = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.box_date_return))
        box_date_return.clear()
        box_date_return.send_keys(regreso_str, Keys.TAB)
        print("Caja de date_return es visible y los datos son ingresados")
    #cambia la cantidad de pasajeros
    @allure.step("Validar cantidad de pasajeros=2")
    def cambiar_cantidad_pasajeros(self):
        try:
            pasajeros = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.number_pasajeros))
            pasajeros.click()
            assert pasajeros.is_displayed(), "El boton no esta verificado"
            print("Boton pasajeros visible y hacemos click")
            btn_mas = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.sumar_pasajeros))
            btn_mas.click()
            time.sleep(2)
            cantidad_pasajeros = pasajeros.get_attribute("value")
            if cantidad_pasajeros != "2":
                print("La cantidad de pasajeros se cambió")
            return True
        except TimeoutException as e:
            print(f"Error: {e}")
    #boton busqueda de vuelos
    @allure.step("Validar la búsqueda ingresando al botón búsqueda de vuelo de un vuelo-ida-regreso")
    def comprar_vuelo(self):
        try:
            btn_comprar = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(self.btn_click_vuelo))
            if btn_comprar.is_enabled():
                btn_comprar.click()
                time.sleep(6)

                print("El botón de comprar de vuelo es visible y hacemos click en él")
            else:
                print("El botón de compra de vuelo no es visible.")
        except TimeoutException:
            print("No se encontró el botón de comprar de vuelo ")
    #selecciona los vuelos de la grilla
    @allure.step("Validar los elementos de vuelo y precios en la grilla de vuelos")
    def precios_vuelos(self):
        try:
            vuelos_ida = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_ida))
            time.sleep(3)
            vuelos_vuelta = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_all_elements_located(self.fecha_vuelta))
            time.sleep(3)

            print("Vuelos de ida:")
            for vuelo in vuelos_ida:
                dia_vuelo_ida = vuelo.text
                precio_vuelo_ida = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_ida > 0:
                    print("Día:", dia_vuelo_ida, "Precio:", precio_vuelo_ida)

            print("Vuelos de regreso:")
            for vuelo in vuelos_vuelta:
                dia_vuelo_regreso = vuelo.text
                precio_vuelo_regreso = self.obtener_precio_vuelo(vuelo)
                if precio_vuelo_regreso > 0:
                    print("Día:", dia_vuelo_regreso,
                        "Precio:", precio_vuelo_regreso)
        except TimeoutException:
            print("Error: No se encontraron vuelos en la fecha seleccionada")
    #metodo para obtener el la fecha y los  precios de los vuelos
    @allure.step("Validar precios")
    def obtener_precio_vuelo(self, vuelo_elemento):
        try:
            precio_elemento = vuelo_elemento.find_element(
                *self.label_locator_monto)
            precio_text = precio_elemento.text.replace(
                '$', '').replace(',', '').strip()
            if precio_text and precio_text.replace('.', '', 1).isdigit():
                return float(precio_text)
            return 0.0
        except Exception as e:
            print(f"Error al obtener el precio del vuelo: {e}")
            return 0.0
    #boton ver vuelo
    @allure.step("Validar la compra ingresando al botón ver vuelos")
    def ver_vuelos(self):
        try:
            print("Esperando que el botón de ver vuelos sea visible")
            ver_vuelo = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located(self.btn_ver_vuelo)
            )

            if ver_vuelo.is_displayed():
                print("El botón de ver vuelos es visible.")
            if ver_vuelo.is_enabled():
                ver_vuelo.click()
                time.sleep(6)
                print("Hicimos click en el botón de ver vuelos.")

                time.sleep(3)
                print(
                    "El botón de ver vuelos es visible y hacemos click en el")
            else:
                print("El botón de ver vuelos no es visible.")
        except TimeoutException:
            print("No se encontró el botón de ver vuelos después de esperar 20 segundos.")
